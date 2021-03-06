import collections
from docx import *
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
import os
import sys
from textwrap import fill


if len(sys.argv)==2:
    Postfix = sys.argv[1]
else:
    Postfix = "A"
    print("Wrong number of arguments")
    print("Continue with A postfix")
    #sys.exit(1)
dirname = os.path.dirname(__file__)
wb = load_workbook( os.path.join(dirname, "../common/responses" + Postfix + ".xlsx") )
ws = wb.active


# Parse labels
wd = Document( os.path.join(dirname, "../common/Anketa.docx") )
questionNumber = 0
Labels = []
for paragraph in wd.paragraphs:
    if paragraph.text.find("Другое") != -1:
        continue
    
    if paragraph.text.find("Вопрос") != -1:
        Labels.append([paragraph.text])
        questionNumber += 1
        continue
    
    if paragraph.text.find("Сколько") != -1:
        column = [cell.value for cell in [column for column in ws.iter_cols(min_col=questionNumber, max_col=questionNumber)][0]][1:]
        Labels[len(Labels) - 1] = Labels[len(Labels) - 1] + list(np.unique(column))[2:]
        continue
    
    if paragraph.style.name == 'List Paragraph':
        # Delete points
        appendedText = paragraph.text
        #print(appendedText)
        if appendedText.rfind('.') != -1:
            appendedText = appendedText[:-1:]
        #print(appendedText)
        #print()
        Labels[len(Labels) - 1].append(appendedText)
        continue
    

# Make plots
def buildMyLovelyPie(column, curChartLabels, myBins = 0):
    if not myBins:
        myBins = range(0, len(curChartLabels)+1 )
    # Preparation
    fig, ax = plt.subplots(figsize=(6, 8), subplot_kw=dict(aspect="equal", anchor = "N"))
    # Count values
    hist = np.histogram(column, bins =  myBins)
    # Extract data and omit zeros
    hist_data = [x for x in hist[0] if x != 0]
    hist_labels = [curChartLabels[i] for i, x in enumerate(hist[0]) if x != 0]
    # Plot 'em all!
    wedges, texts, autotexts = ax.pie( hist_data, autopct = lambda pct: "{:.1f}%".format(pct), pctdistance = 1.2)
    ax.legend(wedges, hist_labels,
              loc="upper center",
              bbox_to_anchor=(0.5, 0))     
    return (fig, ax)


def getColumn(ws, colIndex):
    column = [cell.value for cell in [column for column in ws.iter_cols(min_col=colIndex, max_col=colIndex)][0]][1:]
    return column


questionNumber = 0
sub11 = 0
while (questionNumber < len(Labels)):
    print('a')
    prelabels = 0
    AddictiveString = ""
    questionNumber += 1
    questionNumberCopy = questionNumber
    
    
    # Make a choice!
    if questionNumber == 11:
        decision = '911'
    elif questionNumber == 21:
        decision = 'Make a roll!'
    else:
        decision = 'U no da wae'
    
    
    if decision == 'U no da wae':
        if questionNumber < 11:
            colIndex = questionNumber
        else:
            colIndex = questionNumber + 14
        # Extract data
        column = getColumn(ws, colIndex)
        curChartLabels = [fill(x, 60) for x in list(map(str, Labels[questionNumber-1][1:]))]         
             
    
    elif decision == '911':
        sub11 += 1
        AddictiveString = "."+str(sub11)
        questionNumber -= 1
        if sub11 == 15:
            questionNumber += 1
        colIndex = questionNumber + sub11 - 1
        column = getColumn(ws, colIndex)
        curChartLabels = ["Да", "Нет"]

        
    elif decision == 'Make a roll!':
        colIndex = questionNumber + 14
        column = getColumn(ws, colIndex)
        Categories = [(18, 25), (26, 32), (33, 40), (41, 49), (50, 60)]
        print(column)
        unitedColumn = [j[0] for j in [ [i for i, dia in enumerate(Categories) if (dia[0] <= x and x <= dia[1])] for x in column] if len(j) != 0]
        print(unitedColumn)
        unitedColumn.sort()
        column = unitedColumn
        curChartLabels = [str(x[0])+"-"+str(x[1]) for x in Categories]
        print(column)
        print(curChartLabels)
        

            
    else:
        print('No (?) decision was made')
        sys.exit(2)

    # Universal ending
    
    #### Better be used while creating summary.docx
    #curChartTitle = Labels[questionNumber-1][0]
    #ax.set_title(curChartTitle)
    
    # Bake pies
    if (prelabels):
        fig, ex = buildMyLovelyPie(column, curChartLabels, prelabels+[prelabels[-1]+7])
    else:
        fig, ex = buildMyLovelyPie(column, curChartLabels)
        
    plt.savefig( os.path.join(dirname, "Вопрос "+str(questionNumberCopy)+AddictiveString+Postfix+".png"), bbox_inches = 'tight' )
    plt.close()
       
