from Equation import Expression
from openpyxl import load_workbook
from operator import add, mul
import os
from py_expression_eval import Parser
import re 

# filename = "Казахи-студенты.xlsx"
class RespondentType:

    
    def __init__(self, Distributions):
        self.Distributions = Distributions
    
    
    @classmethod
    def setViaFile(cls, my_filename):
        dirname = os.path.dirname(__file__)
        wb = load_workbook(filename = os.path.join(dirname, './Rules/'+my_filename) )
        ws = wb.active
        
        Distributions = []
        for row in ws:
            Distribution = []
            for enumerator, cell in enumerate(row):
                if cell.value == None:
                    continue
                if str(cell.value).find('SUM') != -1:
                    continue
                if re.search('[а-яА-ЯёЁ]|http', str(cell.value)):
                    continue
                
                #if cell.value == "?":       
                    #print("Hey!!!!", my_filename)
                
                if cell.value == "f:":
                    print("Hello!")
                    i, j = cell.row, cell.col_idx
                    str_expr = str(ws.cell(row = i, column = j + 1).value)
                    print(str_expr)
                    
                    parser = Parser()
                    expr = parser.parse(str_expr)
                    Distribution = Distribution + [float(expr.evaluate({'x':x})) for x in range(ws.cell(row = i, column = j + 2).value, ws.cell(row = i, column = j + 3).value)]
                    print(Distribution)
                    break
                    
                
                Distribution.append(cell.value)
            Distributions.append(Distribution)
        return cls(Distributions)
    
    
    def __add__(self, other):
        Sum = []
        for i in range( len(self.Distributions) ):
            list1 = self.Distributions[i]
            list2 = other.Distributions[i]
            Sum.append( list( map(add, list1, list2) ) )
        return RespondentType(Sum)
    
    def __mul__(self, number):
        Product = []
        for i in range( len(self.Distributions) ):
            row = []
            for item in self.Distributions[i]:
                row.append(number*item)
            Product.append(row)
        return RespondentType(Product)
