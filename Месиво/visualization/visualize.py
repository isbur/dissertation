import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook

wb = load_workbook("responses.xlsx")
ws = wb.active

a = np.histogram( [cell.value for cell in np.array(ws['A'])][1:] )
