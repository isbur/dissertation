import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
from docx import *

wb = load_workbook("responses.xlsx")
ws = wb.active

#a = plt.hist( [cell.value for cell in np.array(ws['E'])][1:] )
#hist = np.histogram([cell.value for cell in np.array(ws['G'])][1:], bins = [0,1,2,3,4,5] )
#a = plt.pie( hist[0] )
#a = plt.pie( hist[0], labels=hist[1][:-1])

wd = Document("Anketa.docx")
count = 0
Labels = []
for paragraph in wd.paragraphs:
    if paragraph.text.find("Другое") != -1:
        continue
    
    if paragraph.text.find("Вопрос") != -1:
                #print(count)
        Labels.append([paragraph.text])
        count += 1
        #print(Labels)
        continue
    
    if paragraph.text.find("Сколько") != -1:
        column = [cell.value for cell in [column for column in ws.iter_cols(min_col=count, max_col=count)][0]][1:]
        Labels[len(Labels) - 1] = Labels[len(Labels) - 1] + list(np.unique(column))[2:]
        continue
    
    
    if paragraph.style.name == 'List Paragraph':
        Labels[len(Labels) - 1].append(paragraph.text)
        continue
    


count = 0
for Label in Labels:
    if count == 3:
        break
    count += 1
    if (count > 2) and (count < 11):
        
        fig, ax = plt.subplots(figsize=(6, 3), subplot_kw=dict(aspect="equal"))
        column = [cell.value for cell in [column for column in ws.iter_cols(min_col=count, max_col=count)][0]][1:]
        hist = np.histogram(column, bins = range(0, len(Labels[count-1][1:])+1 ) )
        wedges, texts = ax.pie( hist[0], labels=Labels[count-1][1:])     
        ax.set_title(Labels[count-1][0])
        plt.savefig("Вопрос "+str(count)+".png")
        plt.close()
        
        
    if count == 11:
        
        heights = []
        for i in range(15):
            column = [cell.value for cell in [column for column in ws.iter_cols(min_col=count+i, max_col=count+i)][0]][1:]
            hist = np.histogram(column, bins = range(0, 3) ) 
            heights.append(hist[0][0])
        
        fig, ax = plt.subplots()
        barplot = plt.bar(x = [x for x in range(1, 16)], height=heights)
        ax.set_title(Labels[count-1][0])
        plt.savefig("Вопрос "+str(count)+".png")
        plt.close()
        
        
    if count > 11:
        
        if count == 21:
            
            column = [cell.value for cell in [column for column in ws.iter_cols(min_col=count+14, max_col=count+14)][0]][1:]
            plt.hist( column, bins = range(len(np.unique(column))+1) )
            plt.savefig("Вопрос "+str(count)+".png")
            plt.close()
            continue
        
        fig, ax = plt.subplots(figsize=(6, 3), subplot_kw=dict(aspect="equal"))
        column = [cell.value for cell in [column for column in ws.iter_cols(min_col=count+14, max_col=count+14)][0]][1:]
        hist = np.histogram(column, bins = range(0, len(Labels[count-1][1:])+1 ) )
        wedges, texts = ax.pie( hist[0], labels=Labels[count-1][1:])     
        ax.set_title(Labels[count-1][0])   
        plt.savefig("Вопрос "+str(count)+".png")
        plt.close()
       
