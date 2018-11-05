import glob
from openpyxl import load_workbook
from pathlib import Path
import xlsxwriter


questions = [
    "1. Страна",
    "2. Национальность",
    "3. Как долго живёте",
    "4. Где",
    "5. Планы",
    "6. Русский",
    "7. Общение на работе",
    "8. Общение вне работы",
    "9. Дети должны",
    "10. Медиа на:",
    "11.1 Взятки на границе",
    "11.2 Взятки в полиции",
    "11.3 Расовая дискриминация",
    "11.4 Религия",
    "11.5 Преступления",
    "11.6 Работа",
    "11.7 Низкая зп",
    "11.8 Работодатель мудак",
    "11.9 Тяжёлый труд",
    "11.10 Жильё",
    "11.11 Медицина",
    "11.12 Юри",
    "11.13 Русский",
    "11.14 Куда с проблемами",
    "11.15 Где работать, жить",
    "12. Есть диаспора?",
    "13. А вы в ней?",
    "14. Первый помощник",
    "15. Гугл работы",
    "16. Гугл жилья",
    "17. НКО",
    "18. Брак с местной",
    "19. РФянин",
    "20. Антимуслим",
    "21. Возраст",
    "22. Образование"
    ]
for filename in glob.glob("*.xlsm"):
    
    if Path(filename).stem in ["responses", "Армяне Типичные Трудовые"]:
        print(Path(filename).stem + " passed .................................")
        continue
    
    readWorkBook = load_workbook(filename)
    readWorkSheet = readWorkBook.active
    writeWorkBook = xlsxwriter.Workbook(str(filename))
    writeWorkBook.add_vba_project('./vbaProject.bin')
    writeWorkSheet = writeWorkBook.add_worksheet()
    writeWorkSheet.insert_button('O3', {'macro':'RandomCellInHColumn',
                                        'caption':'Мне повезёт!',
                                        'width': 200,
                                        'height': 47 })
    writeWorkSheet.write_column(0, 19, questions)
          
    print (filename)
    
    for i, row in enumerate(readWorkSheet.iter_rows()):
        for j, cell in enumerate(row):
            
            if str(cell.value) != 'None':
                #print(i, j, str(cell.value))
                writeWorkSheet.write(i,j, cell.value)
            
    readWorkBook.close()
    writeWorkBook.close()