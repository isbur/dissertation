import glob
from openpyxl import load_workbook
from pathlib import Path
import xlsxwriter

for filename in glob.glob("*.xlsx"):
    
    if Path(filename).stem == "responses":
        continue
    
    readWorkBook = load_workbook(filename)
    readWorkSheet = readWorkBook.active
    readWorkSheet.insert_cols(9)
    filename = Path(filename).stem + ".xlsm"
    writeWorkBook = xlsxwriter.Workbook(str(filename))
    writeWorkBook.add_vba_project('./vbaProject.bin')
    writeWorkSheet = writeWorkBook.add_worksheet()
    writeWorkSheet.insert_button('O3', {'macro':'RandomCellInHColumn',
                                        'caption':'Мне повезёт!',
                                        'width': 200,
                                        'height': 47 })
      
    print (filename)
    
    for i, row in enumerate(readWorkSheet.iter_rows()):
        for j, cell in enumerate(row):
            print(i, j, str(cell.value))
            writeWorkSheet.write(i,j, cell.value)
            
    readWorkBook.close()
    writeWorkBook.close()