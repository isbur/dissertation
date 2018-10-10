from Equation import Expression
from openpyxl import load_workbook
import re 

wb = load_workbook(filename = "Казахи-студенты.xlsx")
ws = wb.active

Distributions = []
for row in ws:
    Distribution = []
    for cell in row:
        if cell.value == None:
            continue
        if str(cell.value).find('SUM') != -1:
            continue
        if re.search('[а-яА-ЯёЁ]', str(cell.value)):
            continue
        
        if cell.value == "f:":
            #print("Hello!")
            i, j = cell.row, cell.col_idx
            fn = Expression(str(ws.cell(row = i, column = j + 1).value), "x")
            #print(fn)
            Distribution = Distribution + [fn(x) for x in range(ws.cell(row = i, column = j + 2).value, ws.cell(row = i, column = j + 3).value)]
            break
            
        
        Distribution.append(cell.value)
    Distributions.append(Distribution)
    