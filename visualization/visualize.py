import collections
from docx import *
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
import os
import sys
from textwrap import fill


if len(sys.argv)==2:
    Postfix = sys.argv[1]
else:
    print("Wrong number of arguments")
    sys.exit(1)
dirname = os.path.dirname(__file__)
wb = load_workbook( os.path.join(dirname, "../common/responses" + Postfix + ".xlsx") )
ws = wb.active


# Parse labels
wd = Document( os.path.join(dirname, "../common/Anketa.docx") )
questionNumber = 0
Labels = []
for paragraph in wd.paragraphs:
    if paragraph.text.find("Другое") != -1:
        continue
    
    if paragraph.text.find("Вопрос") != -1:
        Labels.append([paragraph.text])
        questionNumber += 1
        continue
    
    if paragraph.text.find("Сколько") != -1:
        column = [cell.value for cell in [column for column in ws.iter_cols(min_col=questionNumber, max_col=questionNumber)][0]][1:]
        Labels[len(Labels) - 1] = Labels[len(Labels) - 1] + list(np.unique(column))[2:]
        continue
    
    if paragraph.style.name == 'List Paragraph':
        Labels[len(Labels) - 1].append(paragraph.text)
        continue
    

# Make plots
def buildMyLovelyPie(column, curChartLabels, myBins = 0):
    if not myBins:
        myBins = range(0, len(curChartLabels)+1 )
    # Preparation
    fig, ax = plt.subplots(figsize=(6, 8), subplot_kw=dict(aspect="equal", anchor = "N"))
    # Count values
    hist = np.histogram(column, bins =  myBins)
    # Extract data and omit zeros
    hist_data = [x for x in hist[0] if x != 0]
    hist_labels = [curChartLabels[i] for i, x in enumerate(hist[0]) if x != 0]
    # Plot 'em all!
    wedges, texts, autotexts = ax.pie( hist_data, autopct = lambda pct: "{:.1f}%".format(pct), pctdistance = 1.3)
    ax.legend(wedges, hist_labels,
              loc="center",
              bbox_to_anchor=(0.5, -0.2))     
    return (fig, ax)


def getColumn(ws, colIndex):
    column = [cell.value for cell in [column for column in ws.iter_cols(min_col=colIndex, max_col=colIndex)][0]][1:]
    return column


questionNumber = 0
sub11 = 0
while (questionNumber < len(Labels)):
    
    prelabels = 0
    AddictiveString = ""
    questionNumber += 1
    questionNumberCopy = questionNumber
    
    # Make a choice!
    if questionNumber == 11:
        decision = '911'
    elif questionNumber == 21:
        decision = 'Make a roll!'
    else:
        decision = 'U no da wae'
    
    
    if decision == 'U no da wae':
        if questionNumber < 11:
            colIndex = questionNumber
        else:
            colIndex = questionNumber + 14
        # Extract data
        column = getColumn(ws, colIndex)
        curChartLabels = [fill(x, 60) for x in list(map(str, Labels[questionNumber-1][1:]))]         
             
    
    elif decision == '911':
        sub11 += 1
        AddictiveString = "."+str(sub11)
        questionNumber -= 1
        if sub11 == 15:
            questionNumber += 1
        colIndex = questionNumber + sub11 - 1
        column = getColumn(ws, colIndex)
        curChartLabels = ["Да", "Нет"]

        
    elif decision == 'Make a roll!':
        colIndex = questionNumber + 14
        column = getColumn(ws, colIndex)
        unitedColumn = [x - x % 7 for x in column]
        unitedColumn.sort()
        column = unitedColumn
        prelabels = list(collections.OrderedDict([(x, 1) for x in column]).keys())
        curChartLabels = [str(x)+"-"+str(x+6) for x in prelabels]
        print(column)
        print(curChartLabels)
        

            
    else:
        print('No (?) decision was made')
        sys.exit()

    # Universal ending
    
    #### Better be used while creating summary.docx
    #curChartTitle = Labels[questionNumber-1][0]
    #ax.set_title(curChartTitle)
    
    # Bake pies
    if (prelabels):
        fig, ex = buildMyLovelyPie(column, curChartLabels, prelabels+[prelabels[-1]+7])
    else:
        fig, ex = buildMyLovelyPie(column, curChartLabels)
        
    plt.savefig( os.path.join(dirname, "Вопрос "+str(questionNumberCopy)+AddictiveString+Postfix+".png") )
    plt.close()
       
