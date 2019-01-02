import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
from docx import *

wb = load_workbook("../responses.xlsx")
ws = wb.active

#a = plt.hist( [cell.value for cell in np.array(ws['E'])][1:] )
#hist = np.histogram([cell.value for cell in np.array(ws['G'])][1:], bins = [0,1,2,3,4,5] )
#a = plt.pie( hist[0] )
#a = plt.pie( hist[0], labels=hist[1][:-1])

wd = Document("../Anketa.docx")
count = 0
Labels = []
for paragraph in wd.paragraphs:
    if paragraph.text.find("Другое") != -1:
        continue
    
    if paragraph.text.find("Вопрос") != -1:
                #print(count)
        Labels.append([paragraph.text])
        count += 1
        #print(Labels)
        continue
    
    if paragraph.text.find("Сколько") != -1:
        column = [cell.value for cell in [column for column in ws.iter_cols(min_col=count, max_col=count)][0]][1:]
        Labels[len(Labels) - 1] = Labels[len(Labels) - 1] + list(np.unique(column))[2:]
        continue
    
    
    if paragraph.style.name == 'List Paragraph':
        Labels[len(Labels) - 1].append(paragraph.text)
        continue
    


count = 0
for Label in Labels:
    count += 1
    if count > 11:
        
        if count == 21:
            
            column = [cell.value for cell in [column for column in ws.iter_cols(min_col=count+14, max_col=count+14)][0]][1:]
            uniqueItemsList = list(map(float, np.unique(column)))
            plt.hist( column, bins = uniqueItemsList + [uniqueItemsList[-1]], rwidth = 0.5, align = "left" )
            plt.xticks(uniqueItemsList)
            plt.savefig("Вопрос "+str(count)+".png")
            plt.close()
            continue
       
