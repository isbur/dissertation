import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
from docx import *

wb = load_workbook("responses.xlsx")
ws = wb.active

#a = plt.hist( [cell.value for cell in np.array(ws['E'])][1:] )
#hist = np.histogram([cell.value for cell in np.array(ws['G'])][1:], bins = [0,1,2,3,4,5] )
#a = plt.pie( hist[0] )
#a = plt.pie( hist[0], labels=hist[1][:-1])

wd = Document("Anketa.docx")
count = 0
Labels = []
for paragraph in wd.paragraphs:
    
    if paragraph.text.find("Другое") != -1:
        continue
    
    if paragraph.text.find("Вопрос") != -1:
        count += 1
        #print(count)
        Labels.append([paragraph.text])
        #print(Labels)
        continue
    
    if paragraph.text.find("Сколько") != -1:
        column = [cell.value for cell in [column for column in ws.iter_cols(min_col=7, max_col=7)][0]][1:]
        Labels[len(Labels) - 1] = Labels[len(Labels) - 1] + list(np.unique(column))[2:]
        continue
    
    if paragraph.style.name == 'List Paragraph':
        Labels[len(Labels) - 1].append(paragraph.text)
        continue

